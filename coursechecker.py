import streamlit as st
import pandas as pd
import time
import io
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from openpyxl import load_workbook

# --- PAGE CONFIG ---
st.set_page_config(
    page_title="Course Availability Tool",
    layout="centered"
)

# --- UCHICAGO BRANDING & DYNAMIC THEMING ---
MAROON = "#800000"
LIGHT_GREYSTONE = "#D9D9D9"
GREYSTONE = "#A6A6A6"
DARK_GREYSTONE = "#737373"
BLACK = "#000000"
WHITE = "#FFFFFF"

st.markdown(f"""
    <style>
    /* Global Page Background */
    .stApp {{
        background-color: {WHITE};
    }}
    
    /* Plain Text on White Screen is Black */
    .stMarkdown, p, li, span, div {{
        color: {BLACK};
    }}
    
    /* Massive Single Line Title */
    .title-text {{
        color: {MAROON};
        font-family: 'Crimson Text', serif;
        font-size: 5rem; /* Massive size */
        font-weight: bold;
        white-space: nowrap;
        margin-bottom: 0px;
        padding-bottom: 0px;
        line-height: 1.1;
    }}
    
    /* Credit Text */
    .credit-text {{
        color: {BLACK};
        font-size: 1.2rem;
        margin-top: 0px;
        margin-bottom: 30px;
    }}
    
    /* File Uploader Customization - White Text on Maroon */
    section[data-testid="stFileUploader"] button {{
        background-color: {MAROON} !important;
        color: {WHITE} !important;
    }}
    
    /* Dropdown/Selectbox - White text on Maroon Background */
    div[data-baseweb="select"] > div {{
        background-color: {MAROON} !important;
        color: {WHITE} !important;
    }}
    
    /* Ensure the text inside the dropdown is white */
    div[data-testid="stSelectbox"] div[data-baseweb="select"] * {{
        color: {WHITE} !important;
    }}

    /* Run Button - White text on Gray Background */
    .stButton>button {{
        width: 100%;
        border-radius: 4px;
        height: 3.5em;
        background-color: {DARK_GREYSTONE};
        color: {WHITE} !important;
        font-weight: bold;
        border: none;
        text-transform: uppercase;
        letter-spacing: 1px;
    }}
    
    .stButton>button:hover {{
        background-color: {GREYSTONE};
        color: {WHITE} !important;
    }}

    /* Status Box - White text on Maroon Background */
    .status-box {{
        padding: 20px;
        border-radius: 4px;
        background-color: {MAROON};
        color: {WHITE} !important;
        font-family: 'Proxima Nova', sans-serif;
        margin-bottom: 25px;
        box-shadow: 2px 2px 5px rgba(0,0,0,0.1);
    }}
    
    .status-box b, .status-box span {{
        color: {WHITE} !important;
    }}

    label {{
        color: {BLACK} !important;
        font-weight: bold !important;
    }}
    
    hr {{
        border-top: 1px solid {LIGHT_GREYSTONE};
    }}
    </style>
    """, unsafe_allow_html=True)

# --- HEADER SECTION ---
st.markdown('<p class="title-text">Course Scheduler Checker</p>', unsafe_allow_html=True)
st.markdown('<p class="credit-text">by Ben B.</p>', unsafe_allow_html=True)

st.markdown("""
### Instructions
* **File Upload**: Provide an Excel file (.xlsx) for analysis.
* **Required Formatting**: Ensure Department is in **Column A** and Course Number is in **Column B**.
* **Detection**: The system automatically identifies header rows.
* **Results**: Available courses are recorded with a **'Y' in Column C**.
""")
st.divider()

# --- INPUT AREA ---
uploaded_file = st.file_uploader("Upload course list", type="xlsx")

input_col1, input_col2 = st.columns(2)
with input_col1:
    quarter = st.selectbox("Select Quarter", ["Spring", "Autumn", "Winter", "Summer"])
with input_col2:
    year = st.selectbox("Select Year", [2025, 2026, 2027, 2028])

target_term = f"{quarter} {year}"

st.markdown(" ") 
run_button = st.button("Run Availability Check")
st.divider()

# --- SELENIUM ENGINE ---
def setup_headless_driver():
    options = Options()
    options.add_argument("--headless=new")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")
    options.add_argument("--window-size=1920,1080")
    options.add_argument("user-agent=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36")
    options.binary_location = "/usr/bin/chromium"
    service = Service(executable_path="/usr/bin/chromedriver")
    
    try:
        driver = webdriver.Chrome(service=service, options=options)
        driver.set_page_load_timeout(60)
        return driver
    except Exception:
        alt_service = Service(executable_path="/usr/lib/chromium-browser/chromedriver")
        return webdriver.Chrome(service=alt_service, options=options)

# --- EXECUTION ---
if uploaded_file and run_button:
    file_bytes = uploaded_file.read()
    wb = load_workbook(filename=io.BytesIO(file_bytes))
    ws = wb.active 
    
    first_cell_b = ws.cell(row=1, column=2).value
    start_row = 1
    try:
        if first_cell_b is not None:
            int(str(first_cell_b).strip().split('-')[0])
            start_row = 1
        else:
            start_row = 2 
    except (ValueError, TypeError):
        start_row = 2

    status_card = st.empty()
    progress_bar = st.progress(0)
    found_metric = st.empty()
    
    try:
        status_card.markdown(f'<div class="status-box">Connecting to AIS...</div>', unsafe_allow_html=True)
        driver = setup_headless_driver()
        wait = WebDriverWait(driver, 20)
        
        driver.get("http://coursesearch92.ais.uchicago.edu/psc/prd92guest/EMPLOYEE/HRMS/c/UC_STUDENT_RECORDS_FL.UC_CLASS_SEARCH_FL.GBL")
        
        status_card.markdown(f'<div class="status-box">Setting Search Term: <b>{target_term}</b></div>', unsafe_allow_html=True)
        term_dropdown = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "select[id*='STRM']")))
        Select(term_dropdown).select_by_visible_text(target_term)
        time.sleep(2)
        
        total_rows = ws.max_row
        found_count = 0

        for row in range(start_row, total_rows + 1):
            subj = ws.cell(row=row, column=1).value 
            num = ws.cell(row=row, column=2).value  
            
            if not subj or not num:
                continue
            
            clean_num = str(num).strip().split('-')[0]
            query = f"{str(subj).strip()} {clean_num}"
            
            status_card.markdown(f'<div class="status-box">Scanning: <b>{query}</b> (Row {row} of {total_rows})</div>', unsafe_allow_html=True)
            found_metric.markdown(f"**Current Search Results Found:** `{found_count}`")
            
            search_bar = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "input.ps-edit")))
            search_bar.click()
            search_bar.send_keys(Keys.CONTROL + "a")
            search_bar.send_keys(Keys.DELETE)
            search_bar.send_keys(query + Keys.ENTER)
            
            time.sleep(2)
            page_content = driver.find_element(By.TAG_NAME, "body").text.lower()
            
            if "no results found" not in page_content and clean_num in page_content:
                ws.cell(row=row, column=3).value = "Y"
                found_count += 1
            else:
                ws.cell(row=row, column=3).value = None

            progress_val = (row - start_row + 1) / (total_rows - start_row + 1)
            progress_bar.progress(min(progress_val, 1.0))

        status_card.success(f"Operation complete. {found_count} matches identified.")
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        st.download_button(
            label="Download Verified File",
            data=output,
            file_name=f"Verified_{target_term.replace(' ', '_')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
        driver.quit()

    except Exception as e:
        st.error(f"Internal Error: {str(e)}")
        if 'driver' in locals():
            driver.quit()

elif not uploaded_file and run_button:
    st.warning("Please upload a source file to proceed.")